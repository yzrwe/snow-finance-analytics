"""Build the quarterly earnings pack: a formatted multi-tab Excel export.

Tabs:
    1. Summary       — headline KPIs for the latest quarter
    2. Revenue Trend — full quarterly series with growth calcs
    3. Health        — NRR / RPO / $1M+ customer trends

Usage:
    python exports/build_earnings_pack.py [--quarter "Q1 FY27"]
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "snow_quarterly_metrics.csv"

BLUE = "29B5E8"
DARK = "0B5394"
HEADER_FILL = PatternFill("solid", fgColor=DARK)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color=DARK)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(quarter: str | None) -> pd.DataFrame:
    df = pd.read_csv(DATA, parse_dates=["quarter_end"]).sort_values("quarter_end")
    df["yoy_growth"] = df["product_revenue_m"] / df["product_revenue_m"].shift(4) - 1
    df["qoq_growth"] = df["product_revenue_m"] / df["product_revenue_m"].shift(1) - 1
    df["seq_add_m"] = df["product_revenue_m"].diff()
    if quarter:
        if quarter not in set(df["fiscal_quarter"]):
            raise SystemExit(f"Unknown quarter {quarter!r}. "
                             f"Available: {', '.join(df['fiscal_quarter'])}")
        idx = df.index[df["fiscal_quarter"] == quarter][0]
        df = df.loc[:idx]
    return df


def style_header(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def autofit(ws) -> None:
    for col_cells in ws.columns:
        width = max(len(str(c.value)) for c in col_cells if c.value is not None)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width + 3


def write_frame(ws, df: pd.DataFrame, start_row: int, pct_cols: set[str],
                money_cols: set[str]) -> None:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True),
                                start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = BORDER
            col_name = df.columns[c_idx - 1]
            if r_idx > start_row:
                if col_name in pct_cols:
                    cell.number_format = "0.0%"
                elif col_name in money_cols:
                    cell.number_format = "#,##0.0"
    style_header(ws, start_row, len(df.columns))


def build(quarter: str | None, out: Path) -> Path:
    df = load(quarter)
    latest, prior = df.iloc[-1], df.iloc[-2]
    wb = Workbook()

    # --- Tab 1: Summary -------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Quarterly Earnings Pack — {latest['fiscal_quarter']}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Quarter ended {latest['quarter_end']:%B %d, %Y}"
    rows = [
        ("Metric", "Value", "Δ vs prior quarter"),
        ("Product revenue ($M)", round(latest["product_revenue_m"], 1),
         f"{latest['qoq_growth']:+.1%}"),
        ("YoY growth", f"{latest['yoy_growth']:.1%}",
         f"{(latest['yoy_growth'] - prior['yoy_growth']) * 100:+.1f} pts"),
        ("Sequential dollar add ($M)", round(latest["seq_add_m"], 1), "—"),
        ("Net revenue retention", f"{latest['nrr_pct']:.0f}%",
         f"{latest['nrr_pct'] - prior['nrr_pct']:+.0f} pts"),
        ("RPO ($B)", round(latest["rpo_b"], 2),
         f"{latest['rpo_b'] / prior['rpo_b'] - 1:+.1%}"),
        ("$1M+ customers", int(latest["customers_1m_plus"]),
         f"{int(latest['customers_1m_plus'] - prior['customers_1m_plus']):+d}"),
    ]
    for r, row in enumerate(rows, start=4):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
    style_header(ws, 4, 3)
    autofit(ws)

    # --- Tab 2: Revenue Trend --------------------------------------------
    ws2 = wb.create_sheet("Revenue Trend")
    trend = df[["fiscal_quarter", "product_revenue_m", "seq_add_m",
                "qoq_growth", "yoy_growth"]].rename(columns={
        "fiscal_quarter": "Quarter", "product_revenue_m": "Product Rev ($M)",
        "seq_add_m": "Seq Add ($M)", "qoq_growth": "QoQ", "yoy_growth": "YoY"})
    write_frame(ws2, trend, 1, pct_cols={"QoQ", "YoY"},
                money_cols={"Product Rev ($M)", "Seq Add ($M)"})
    autofit(ws2)

    # --- Tab 3: Health ----------------------------------------------------
    ws3 = wb.create_sheet("Health")
    health = df[["fiscal_quarter", "nrr_pct", "rpo_b", "customers_1m_plus"]].rename(
        columns={"fiscal_quarter": "Quarter", "nrr_pct": "NRR (%)",
                 "rpo_b": "RPO ($B)", "customers_1m_plus": "$1M+ Customers"})
    write_frame(ws3, health, 1, pct_cols=set(), money_cols={"RPO ($B)"})
    autofit(ws3)

    wb.save(out)
    return out


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--quarter", default=None, help='e.g. "Q1 FY27"')
    parser.add_argument("--out", default=None)
    args = parser.parse_args()
    label = (args.quarter or "latest").replace(" ", "_")
    out = Path(args.out) if args.out else ROOT / "exports" / f"earnings_pack_{label}.xlsx"
    path = build(args.quarter, out)
    print(f"✓ Earnings pack written to {path}")
